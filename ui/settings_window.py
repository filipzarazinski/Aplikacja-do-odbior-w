"""
ui/settings_window.py
Okno ustawień aplikacji – zakładki: Kolumny, Baza SIM.
"""
import io
import logging
import os
from typing import Optional

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QTabWidget, QWidget,
    QLabel, QLineEdit, QPushButton, QFileDialog,
    QListWidget, QListWidgetItem, QAbstractItemView,
    QGroupBox, QCheckBox, QColorDialog, QRadioButton, QButtonGroup,
    QMessageBox, QDialogButtonBox, QProgressDialog, QApplication,
)
from PySide6.QtCore import Qt, QThread, Signal, Slot
from PySide6.QtGui import QColor

from database.db_manager import DatabaseManager
from ui.widgets.montaz_tab import _ID_TO_MODEL

logger = logging.getLogger(__name__)



# ── Worker thread do synchronizacji ───────────────────────────────────────────

class SyncWorker(QThread):
    progress = Signal(str)
    finished = Signal(bool, str)   # ok, message

    def __init__(self, source: str, is_url: bool):
        super().__init__()
        self._source = source
        self._is_url = is_url

    def run(self):
        try:
            if self._is_url:
                content = self._download(self._source)
            else:
                with open(self._source, "rb") as f:
                    content = f.read()

            cards = self._parse_excel(content)
            if not cards:
                self.finished.emit(False, "Nie znaleziono danych (puste arkusze?)")
                return

            db = DatabaseManager.instance()
            count = db.bulk_upsert_sim_cards(cards)
            from datetime import datetime
            db.set_setting("sim_last_sync", datetime.now().strftime("%Y-%m-%d %H:%M"))
            db.set_setting("sim_source", self._source)
            self.finished.emit(True, f"Zsynchronizowano {count} kart SIM.")

        except MissingLibraryError as e:
            self.finished.emit(False, str(e))
        except Exception as e:
            logger.error("Sync error", exc_info=True)
            self.finished.emit(False, f"Błąd: {e}")

    def _download(self, url: str) -> bytes:
        try:
            import requests
        except ImportError:
            raise MissingLibraryError(
                "Brak biblioteki 'requests'.\nUruchom: pip install requests openpyxl"
            )
        self.progress.emit("Pobieranie pliku z SharePoint…")
        # Próba pobrania z parametrem download=1
        dl_url = url.split("&action=")[0] + "&download=1"
        r = requests.get(dl_url, timeout=30)
        if r.status_code == 401:
            raise RuntimeError(
                "Brak dostępu (401). Plik jest chroniony.\n"
                "Pobierz plik ręcznie przez przeglądarkę i wskaż jego lokalizację."
            )
        if r.status_code == 403:
            raise RuntimeError(
                "Brak uprawnień (403). Zaloguj się lub pobierz plik ręcznie."
            )
        r.raise_for_status()
        return r.content

    def _parse_excel(self, content: bytes) -> list[tuple[str, str]]:
        try:
            import openpyxl
        except ImportError:
            raise MissingLibraryError(
                "Brak biblioteki 'openpyxl'.\nUruchom: pip install openpyxl"
            )
        self.progress.emit("Parsowanie pliku Excel…")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        cards = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                # Sprawdź czy to nagłówek
                first = str(row[0] or "").strip().lower()
                if not first.lstrip("+").isdigit():
                    continue  # pomiń nagłówek
            sim = str(row[0] or "").strip()
            ccid = str(row[1] or "").strip()
            if sim and ccid and sim.lower() not in ("none", "sim", "numer sim"):
                cards.append((sim, ccid))
        wb.close()
        return cards


class MissingLibraryError(Exception):
    pass


def _detect_device_model(device_id: str) -> str:
    """Wykrywa model urządzenia na podstawie ID — ta sama logika co w formularzu."""
    t = (device_id or "").strip()
    for pattern, model in _ID_TO_MODEL:
        if pattern.match(t):
            return model
    return ""


# ── Worker thread do importu Odbiory.xlsb ─────────────────────────────────────

class OdbioryImportWorker(QThread):
    progress = Signal(str)
    row_done = Signal(int, int)                             # current, total
    finished = Signal(bool, str, int)                        # ok, msg, errors

    def __init__(self, path: str, overwrite: bool = False):
        super().__init__()
        self._path = path
        self._overwrite = overwrite

    def run(self):
        import importlib.util
        if importlib.util.find_spec("pyxlsb") is None:
            self.finished.emit(False, "Brak biblioteki 'pyxlsb'. Uruchom: pip install pyxlsb", 0)
            return
        try:
            self._do_import()
        except Exception as e:
            logger.error("Import Odbiory error", exc_info=True)
            self.finished.emit(False, f"Błąd: {e}", 0)

    def _do_import(self):
        import pyxlsb

        self.progress.emit("Otwieranie pliku…")
        header_row: list[str] | None = None
        rows: list[list] = []

        with pyxlsb.open_workbook(self._path) as wb:
            with wb.get_sheet("Odbiory") as ws:
                for i, row in enumerate(ws.rows()):
                    if i == 0:
                        first_v = row[0].v if row else None
                        if isinstance(first_v, str) and first_v.strip():
                            header_row = [str(c.v or "").strip() for c in row]
                        else:
                            rows.append([c.v for c in row])
                    else:
                        rows.append([c.v for c in row])

        col_map = self._build_col_map(header_row) if header_row else {}

        total = len(rows)
        self.progress.emit(f"Wczytano {total} wierszy, importowanie…")

        db = DatabaseManager.instance()
        inserted = 0
        empty = 0
        errors = 0
        COMMIT_EVERY = 100

        for idx, cells in enumerate(rows):
            self.row_done.emit(idx + 1, total)
            if not any(cells):
                empty += 1
                continue
            try:
                rec = self._cells_to_record(cells, col_map)
                if rec is None:
                    empty += 1
                    continue
                db.insert_record_no_commit(rec)
                inserted += 1
                if inserted % COMMIT_EVERY == 0:
                    db.commit()
            except Exception as e:
                logger.warning(f"Row {idx + 2} import error: {e}")
                errors += 1

        db.commit()
        parts = [f"Dodano: {inserted}", f"Puste/bez daty: {empty}"]
        if errors:
            parts.append(f"Błędy: {errors}")
        msg = "Import zakończony.  " + "   |   ".join(parts)
        self.finished.emit(True, msg, errors)

    @staticmethod
    def _build_col_map(headers: list[str]) -> dict[str, int]:
        """Build field-key → column-index map from header row."""
        import unicodedata

        _EXTRA = str.maketrans('łŁ', 'll')   # NFD doesn't decompose ł

        def norm(s: str) -> str:
            s = s.lower().strip().translate(_EXTRA)
            s = unicodedata.normalize('NFD', s)
            s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
            return ''.join(c for c in s if c.isalnum())

        # Ordered: more specific/longer substrings first
        PATTERNS = [
            ("sonda1id",          "probe1_id"),
            ("sonda1poj",         "probe1_cap"),
            ("sonda1dl",          "probe1_len"),
            ("sonda2id",          "probe2_id"),
            ("sonda2poj",         "probe2_cap"),
            ("sonda2dl",          "probe2_len"),
            ("ktoryzbiornik",     "right_tank_probe"),
            ("komentarzprywatny", "private_comment"),
            ("czasdyzurow",       "duty_time"),
            ("typrejestratora",   "recorder_type"),
            ("rodzajpojazdu",     "vehicle_type"),
            ("canconfigname",     "_skip_can_config_name"),  # absorb before "canconfig"
            ("canconfig",         "can_config"),
            ("dinconfig",         "din_config"),
            ("additionalconfig",  "additional_config"),
            ("numerrejestracyjn", "license_plate"),
            ("numerboczn",        "side_number"),
            ("gdzierejestrator",  "recorder_location"),
            ("markaimodel",       "vehicle_brand"),
            ("markamodel",        "vehicle_brand"),
            ("firmwaretacho",     "firmware_tacho"),
            ("modelurzadzen",     "device_model"),
            ("montazserwis",      "record_type"),
            ("czynnosciwykon",    "comment"),   # "Czynności wykonane przez montera"
            ("przebieg",          "mileage"),
            ("immorfid",          "immo_rfid"),
            ("immo",              "immo_rfid"),
            ("tablet",            "tablet"),
            ("firma",             "company_name"),
            ("flota",             "fleet_name"),
        ]
        EXACT = [
            ("data",    "date"),
            ("czas",    "time"),
            ("sim",     "sim_number"),
            ("id",      "device_id"),
            ("monter",  "technician"),  # exact to avoid matching "...przez montera"
        ]

        result: dict[str, int] = {}
        assigned: set[int] = set()
        normalized = [(idx, norm(h)) for idx, h in enumerate(headers)]

        for col_idx, hn in normalized:
            if not hn or col_idx in assigned:
                continue
            for pattern, field_key in PATTERNS:
                if field_key not in result and pattern in hn:
                    result[field_key] = col_idx
                    assigned.add(col_idx)
                    break

        for col_idx, hn in normalized:
            if not hn or col_idx in assigned:
                continue
            for pattern, field_key in EXACT:
                if field_key not in result and hn == pattern:
                    result[field_key] = col_idx
                    assigned.add(col_idx)
                    break

        # "Numer SIM" → "numerism" (removes space) which contains "sim"
        if "sim_number" not in result:
            for col_idx, hn in normalized:
                if col_idx not in assigned and "sim" in hn:
                    result["sim_number"] = col_idx
                    break

        return result

    @staticmethod
    def _cells_to_record(cells, col_map: dict):
        from datetime import datetime as _dt, timedelta
        import json
        from database.models import ServiceRecord, DinChannel
        from config import CAN_JSON_KEYS, CAN_CONNECTION_TRUCK, CAN_CONNECTION_CAR

        def _get(field_key: str, fallback_idx: int, default=None):
            idx = col_map.get(field_key, fallback_idx) if col_map else fallback_idx
            if idx is None or idx >= len(cells):
                return default
            return cells[idx]

        def _s(field_key: str, fallback_idx: int) -> str:
            v = _get(field_key, fallback_idx)
            return "" if v is None else str(v).strip()

        # --- Date ---
        date_v = _get("date", 0)
        service_date = ""
        if isinstance(date_v, (int, float)) and date_v > 0:
            service_date = (_dt(1899, 12, 30) + timedelta(days=int(date_v))).strftime("%Y-%m-%d")
        elif isinstance(date_v, str):
            sv = date_v.strip()
            if "." in sv:
                parts = sv.split(".")
                if len(parts) == 3:
                    service_date = f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
                else:
                    service_date = sv
            else:
                service_date = sv
        if not service_date:
            return None

        # --- Time ---
        time_v = _get("time", 15)
        service_hour, service_minute = 0, 0
        if isinstance(time_v, (int, float)) and time_v > 0:
            total_min = round(time_v * 24 * 60)
            service_hour = (total_min // 60) % 24
            service_minute = total_min % 60

        # --- Monter ---
        technician_name = _s("technician", 14)

        # --- Side number (may be float) ---
        side_v = _get("side_number", 6)
        if isinstance(side_v, float) and side_v == int(side_v):
            side_number = str(int(side_v))
        else:
            side_number = "" if side_v is None else str(side_v).strip()

        # --- Mileage ---
        mileage_v = _get("mileage", 19)
        mileage = None
        if isinstance(mileage_v, (int, float)) and mileage_v > 0:
            mileage = int(mileage_v)
        elif isinstance(mileage_v, str):
            try:
                mileage = int(float(mileage_v.replace(",", ".")))
            except (ValueError, TypeError):
                pass

        # --- Duty time ---
        duty_time_min = _s("duty_time", 13) or None

        # --- Probe float helper ---
        def probe_float(field_key: str, fallback_idx: int):
            v = _get(field_key, fallback_idx)
            if isinstance(v, (int, float)):
                return float(v) if v else None
            if isinstance(v, str):
                try:
                    return float(v.replace(",", "."))
                except (ValueError, TypeError):
                    pass
            return None

        # --- CAN config ---
        can_active = False
        can_checkboxes = [False] * 8
        can_vehicle_type = ""
        can_cfg: dict = {}
        can_raw = _s("can_config", 31)
        if can_raw:
            try:
                can_cfg = json.loads(can_raw)
                can_active = str(can_cfg.get("isCan", "")).lower() == "true"
                for i, key in enumerate(CAN_JSON_KEYS):
                    can_checkboxes[i] = str(can_cfg.get(key, "")).lower() == "tak"
                conn = can_cfg.get("canConnection", "")
                if conn == CAN_CONNECTION_TRUCK:
                    can_vehicle_type = "Ciężarowy"
                elif conn == CAN_CONNECTION_CAR:
                    can_vehicle_type = "Osobowy"
            except (json.JSONDecodeError, TypeError):
                pass

        # --- DIN config ---
        din1, din2, din3 = DinChannel(), DinChannel(), DinChannel()
        din_cfg: dict = {}
        din_raw = _s("din_config", 33)
        if din_raw:
            try:
                din_d = json.loads(din_raw)
                if din_d:
                    if any(k.startswith("din") and k[3:].isdigit() for k in din_d):
                        din_cfg = din_d
                    else:
                        next_idx = 1
                        if "webasto" in din_d:
                            props = din_d["webasto"]
                            bit = str(props.get("bit", props.get("din", "")))
                            din_cfg["din1"] = {"nazwa": "webasto", "bit": bit, "stan": props.get("stan", "wysoki")}
                            next_idx = 2
                        for device_name, props in din_d.items():
                            if device_name == "webasto":
                                continue
                            bit = str(props.get("bit", props.get("din", "")))
                            din_cfg[f"din{next_idx}"] = {
                                "nazwa": device_name, "bit": bit,
                                "stan": props.get("stan", "wysoki"), "sn": str(props.get("sn", ""))
                            }
                            next_idx += 1

                    din_channels = []
                    for key in sorted((k for k in din_cfg if k.startswith("din") and k[3:].isdigit()), key=lambda x: int(x[3:])):
                        d = din_cfg[key]
                        bit = str(d.get("bit", ""))
                        stan = d.get("stan", "")
                        din_channels.append(DinChannel(
                            function=d.get("nazwa", ""),
                            din_type=bit,
                            level_low=(stan == "niski"),
                            level_high=(stan == "wysoki"),
                            serial_number=str(d.get("sn", "")),
                        ))
                    if din_channels:
                        din1 = din_channels[0]
                    if len(din_channels) > 1:
                        din2 = din_channels[1]
                    if len(din_channels) > 2:
                        din3 = din_channels[2]
            except (json.JSONDecodeError, TypeError):
                pass

        # --- Tablet: "True|sn|power" ---
        has_tablet, tablet_sn, has_power = False, "", False
        tablet_raw = _s("tablet", 41)
        if tablet_raw:
            parts = tablet_raw.split("|")
            has_tablet = parts[0].lower() == "true"
            tablet_sn = parts[1].strip() if len(parts) > 1 else ""
            has_power = (parts[2].strip() == "1") if len(parts) > 2 else False

        # --- Immo|RFID ---
        has_immo, has_rfid = False, False
        immo_raw = _s("immo_rfid", 42)
        if "|" in immo_raw:
            parts = immo_raw.split("|")
            has_immo = parts[0].strip() == "1"
            has_rfid = parts[1].strip() == "1" if len(parts) > 1 else False

        # --- Additional config ---
        add_cfg: dict = {}
        add_raw = _s("additional_config", 34)
        if add_raw:
            try:
                add_cfg = json.loads(add_raw)
            except (json.JSONDecodeError, TypeError):
                pass
        add_cfg["immo"] = "1" if has_immo else ""
        add_cfg["rfid"] = "1" if has_rfid else ""

        is_weekend = False
        if service_date:
            try:
                dt_obj = _dt.strptime(service_date, "%Y-%m-%d")
                if dt_obj.weekday() >= 5:
                    is_weekend = True
            except ValueError:
                pass

        dyzur = is_weekend or service_hour >= 15 or service_hour < 6 or (service_hour == 6 and service_minute <= 55)

        config_json: dict = {
            "canConfig": can_cfg,
            "dinConfig": din_cfg,
            "additionalConfig": add_cfg,
            "komentarzPrywatny": _s("private_comment", 44),
            "odebrane": True,
            "dyzurZaznaczony": dyzur,
        }

        rt = _s("record_type", 2)
        return ServiceRecord(
            record_type=rt if rt else " ",
            service_date=service_date,
            service_hour=service_hour,
            service_minute=service_minute,
            company_name=_s("company_name", 3),
            fleet_name=_s("fleet_name", 4),
            license_plate=_s("license_plate", 5),
            side_number=side_number,
            vehicle_brand=_s("vehicle_brand", 9),
            vehicle_type=_s("vehicle_type", 28),
            device_id=_s("device_id", 7),
            sim_number=_s("sim_number", 8),
            device_model=_detect_device_model(_s("device_id", 7)),
            firmware_tacho=_s("firmware_tacho", 17),
            recorder_location=_s("recorder_location", 18),
            mileage=mileage,
            probe1_id=_s("probe1_id", 20),
            probe1_capacity=probe_float("probe1_cap", 21),
            probe1_length=probe_float("probe1_len", 22),
            probe2_id=_s("probe2_id", 23),
            probe2_capacity=probe_float("probe2_cap", 24),
            probe2_length=probe_float("probe2_len", 25),
            right_tank_probe=_s("right_tank_probe", 26),
            can_active=can_active,
            can_checkboxes=can_checkboxes,
            can_vehicle_type=can_vehicle_type,
            din1=din1,
            din2=din2,
            din3=din3,
            has_rfid=has_rfid,
            has_immo=has_immo,
            has_tablet=has_tablet,
            tablet_sn=tablet_sn,
            has_power=has_power,
            config_json=config_json,
            technician_name=technician_name,
            comment=_s("comment", 11),
            duty_time_min=duty_time_min,
        )


# ── Główny dialog ──────────────────────────────────────────────────────────────

class SettingsWindow(QDialog):

    columns_changed = Signal(set, list)   # visible_set, column_order

    def __init__(self, all_columns: list, column_order: list, visible_set: set, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ustawienia")
        self.setModal(True)
        self.resize(600, 620)

        self._db = DatabaseManager.instance()
        self._is_light = self._db.get_setting("theme_mode", "dark") == "light"
        
        self._all_columns = all_columns
        self._column_order = list(column_order)
        self._visible_set = set(visible_set)
        self._worker: Optional[SyncWorker] = None
        self._import_worker: Optional[OdbioryImportWorker] = None

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        tabs = QTabWidget()
        tabs.setDocumentMode(True)
        tabs.addTab(self._build_general_tab(),         "  Ogólne  ")
        tabs.addTab(self._build_columns_tab(),         "  Kolumny  ")
        tabs.addTab(self._build_dicts_tab(),           "  Słowniki  ")
        tabs.addTab(self._build_colors_tab(),          "  Kolory  ")
        tabs.addTab(self._build_odbiory_import_tab(),  "  Import z Excel  ")
        tabs.addTab(self._build_about_tab(),           "  O aplikacji  ")
        root.addWidget(tabs, 1)

        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(12, 8, 12, 10)
        btn_close = QPushButton("Zamknij")
        btn_close.setFixedSize(90, 28)
        btn_close.clicked.connect(self.accept)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        root.addLayout(btn_row)

    # ---------------------------------------------------------------- Dicts tab (kontener)

    def _build_dicts_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        inner = QTabWidget()
        inner.setDocumentMode(True)
        inner.addTab(self._build_sim_tab(),              "  Baza SIM  ")
        inner.addTab(self._build_zbiorcze_tab(),         "  Słowniki / Importy  ")
        inner.addTab(self._build_bulk_import_widget(),   "  Zbiorczy import słowników  ")
        inner.currentChanged.connect(
            lambda idx: self._sim_dict_tab.ensure_loaded() if idx == 0 else None
        )
        lay.addWidget(inner, 1)
        return w

    # ---------------------------------------------------------------- SIM tab

    def _build_sim_tab(self) -> QWidget:
        from PySide6.QtWidgets import QSplitter
        from PySide6.QtCore import Qt as _Qt
        from ui.dict_tab import DictTab

        w = QWidget()
        outer = QVBoxLayout(w)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        splitter = QSplitter(_Qt.Vertical)
        splitter.setChildrenCollapsible(False)

        # ── Panel górny: synchronizacja + status ────────────────────────────
        top = QWidget()
        lay = QVBoxLayout(top)
        lay.setContentsMargins(8, 8, 8, 4)
        lay.setSpacing(5)

        # --- Synchronizacja z pliku ---
        file_row = QHBoxLayout()
        file_row.setSpacing(4)
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("Ścieżka do pliku .xlsx / .xlsm …")
        saved_file = self._db.get_setting("sim_source_file", "")
        self._file_edit.setText(saved_file)
        file_row.addWidget(self._file_edit, 1)

        btn_browse = QPushButton("…")
        btn_browse.setFixedWidth(28)
        btn_browse.clicked.connect(self._on_browse)
        file_row.addWidget(btn_browse)

        btn_sync_file = QPushButton("📥  Wczytaj z pliku")
        btn_sync_file.setObjectName("btn_primary")
        btn_sync_file.clicked.connect(self._on_sync_file)
        file_row.addWidget(btn_sync_file)
        lay.addLayout(file_row)

        # --- Status ---
        status_row = QHBoxLayout()
        status_row.setSpacing(4)
        last_sync = self._db.get_setting("sim_last_sync", "–")
        count = self._db.get_sim_cards_count()
        self._lbl_last  = QLabel(f"Ostatnia sync:  {last_sync}")
        self._lbl_count = QLabel(f"Kart SIM:  {count}")
        self._progress_lbl = QLabel("")
        for lbl in (self._lbl_last, self._lbl_count, self._progress_lbl):
            lbl.setStyleSheet("font-size: 8.5pt;")
        self._progress_lbl.setStyleSheet("color: #64748b; font-size: 8.5pt;")

        btn_clear_sim = QPushButton("🗑  Wyczyść")
        btn_clear_sim.setToolTip("Usuwa wszystkie karty SIM z bazy danych")
        btn_clear_sim.clicked.connect(self._on_clear_sim)

        status_row.addWidget(self._lbl_last)
        status_row.addWidget(QLabel("|"))
        status_row.addWidget(self._lbl_count)
        status_row.addWidget(self._progress_lbl)
        status_row.addStretch()
        status_row.addWidget(btn_clear_sim)
        lay.addLayout(status_row)

        note = QLabel("ℹ  Kol. A = SIM,  kol. B = CCID  (wiersz 1 = nagłówek – pomijany).  pip install openpyxl")
        note.setStyleSheet("color: #64748b; font-size: 8pt;")
        lay.addWidget(note)

        splitter.addWidget(top)

        # ── Panel dolny: tabela kart SIM ────────────────────────────────────
        def _sim_excel_parser(ws):
            result = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    first = str(row[0] or "").strip().lower()
                    if not first.lstrip("+").isdigit():
                        continue
                sim  = str(row[0] or "").strip()
                ccid = str(row[1] or "").strip() if len(row) > 1 else ""
                if sim and ccid and sim.lower() not in ("none", "sim"):
                    result.append([sim, ccid])
            return result

        self._sim_dict_tab = DictTab(
            headers=["Numer SIM", "CCID", "Data synchronizacji"],
            loader=self._db.get_all_sim_cards,
            saver=lambda vals: self._db.upsert_sim_card(vals[0], vals[1]),
            deleter=self._db.delete_sim_card_by_id,
            excel_sheet="SIM",
            excel_parser=_sim_excel_parser,
            add_labels=["Numer SIM", "CCID"],
            commit_fn=self._db.commit,
            lazy=True,
        )
        splitter.addWidget(self._sim_dict_tab)

        splitter.setSizes([110, 500])
        outer.addWidget(splitter, 1)
        return w

    # ---------------------------------------------------------------- Firmy tab

    def _build_firmy_tab(self) -> QWidget:
        from ui.dict_tab import DictTab

        def saver(vals):
            self._db.upsert_company_with_fleet(vals[0], vals[1] if len(vals) > 1 else "")

        def excel_parser(ws):
            """Kolumny = floty, wiersze = firmy."""
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            result = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                for col_idx, fleet in enumerate(headers):
                    if not fleet:
                        continue
                    company = str(row[col_idx] or "").strip() if col_idx < len(row) else ""
                    if company and company.lower() not in ("none", ""):
                        result.append([company, fleet])
            return result

        tab = DictTab(
            headers=["Nazwa firmy", "Flota"],
            loader=self._db.get_all_companies_with_fleet,
            saver=saver,
            deleter=self._db.delete_company_by_id,
            excel_sheet="Firmy",
            excel_parser=excel_parser,
            add_labels=["Nazwa firmy", "Flota"],
            commit_fn=self._db.commit,
        )
        return tab

    # ---------------------------------------------------------------- Pojazdy tab

    def _build_pojazdy_tab(self) -> QWidget:
        from ui.dict_tab import DictTab

        def saver(vals):
            self._db.upsert_vehicle_model(vals[0], vals[1] if len(vals) > 1 else "")

        def excel_parser(ws):
            result = []
            for row in ws.iter_rows(values_only=True):
                model = str(row[0] or "").strip()
                vtype = str(row[1] or "").strip() if len(row) > 1 else ""
                if model and model.lower() not in ("none", "marka", "model"):
                    result.append([model, vtype])
            return result

        return DictTab(
            headers=["Marka i model", "Typ pojazdu"],
            loader=self._db.get_all_vehicle_models,
            saver=saver,
            deleter=self._db.delete_vehicle_model_by_id,
            excel_sheet="Model_Typ",
            excel_parser=excel_parser,
            add_labels=["Marka i model", "Typ pojazdu"],
            commit_fn=self._db.commit,
        )

    # ---------------------------------------------------------------- Monterzy tab

    def _build_monterzy_tab(self) -> QWidget:
        from ui.dict_tab import DictTab
        from database.models import Technician

        def saver(vals):
            self._db.upsert_technician(Technician(full_name=vals[0]))

        def loader():
            return [(t.id, t.full_name) for t in self._db.get_all_technicians(active_only=False)]

        def excel_parser(ws):
            result = []
            for row in ws.iter_rows(values_only=True):
                name = str(row[0] or "").strip()
                if name and name.lower() not in ("none", "monter"):
                    result.append([name])
            return result

        return DictTab(
            headers=["Imie i nazwisko"],
            loader=loader,
            saver=saver,
            deleter=self._db.delete_technician_by_id,
            excel_sheet="Monterzy",
            excel_parser=excel_parser,
            add_labels=["Imie i nazwisko"],
            commit_fn=self._db.commit,
        )

    # ---------------------------------------------------------------- Lokalizacje tab

    def _build_loki_tab(self) -> QWidget:
        from ui.dict_tab import DictTab

        def saver(vals):
            self._db.upsert_recorder_location(vals[0])

        def excel_parser(ws):
            result = []
            for row in ws.iter_rows(values_only=True):
                loc = str(row[0] or "").strip()
                if loc and loc.lower() not in ("none", "lokalizacja"):
                    result.append([loc])
            return result

        return DictTab(
            headers=["Lokalizacja"],
            loader=self._db.get_all_recorder_locations,
            saver=saver,
            deleter=self._db.delete_recorder_location_by_id,
            excel_sheet="GdzieRejestrator",
            excel_parser=excel_parser,
            add_labels=["Lokalizacja"],
            commit_fn=self._db.commit,
        )

    # ---------------------------------------------------------------- Model urządzenia tab

    def _build_device_models_tab(self) -> QWidget:
        from ui.dict_tab import DictTab

        def saver(vals):
            self._db.upsert_device_model(vals[0])

        def excel_parser(ws):
            result = []
            for row in ws.iter_rows(values_only=True):
                name = str(row[0] or "").strip()
                if name and name.lower() not in ("none", "model", "model urządzenia"):
                    result.append([name])
            return result

        return DictTab(
            headers=["Model urządzenia"],
            loader=self._db.get_all_device_models,
            saver=saver,
            deleter=self._db.delete_device_model_by_id,
            excel_sheet="ModelUrzadzenia",
            excel_parser=excel_parser,
            add_labels=["Model urządzenia"],
            commit_fn=self._db.commit,
        )

    # ---------------------------------------------------------------- Urządzenia dodatkowe tab

    def _build_extra_devices_tab(self) -> QWidget:
        from ui.dict_tab import DictTab

        def saver(vals):
            self._db.upsert_extra_device(
                vals[1] if len(vals) > 1 else "",  # fleet_name
                vals[0],                            # device_name
            )

        def excel_parser(ws):
            """Kolumny = floty (nagłówek), wiersze = urządzenia dodatkowe."""
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            result = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                for col_idx, fleet in enumerate(headers):
                    if not fleet:
                        continue
                    device = str(row[col_idx] or "").strip() if col_idx < len(row) else ""
                    if device and device.lower() not in ("none", ""):
                        result.append([device, fleet])
            return result

        def loader():
            # Zwraca (id, device_name, fleet_name) — zamiana kolejności dla czytelności w tabeli
            return [(r[0], r[2], r[1]) for r in self._db.get_all_extra_devices()]

        return DictTab(
            headers=["Urządzenie dodatkowe", "Flota"],
            loader=loader,
            saver=saver,
            deleter=self._db.delete_extra_device_by_id,
            excel_sheet="UrzadzeniaDodatkowe",
            excel_parser=excel_parser,
            add_labels=["Urządzenie dodatkowe", "Flota"],
            commit_fn=self._db.commit,
        )

    # ---------------------------------------------------------------- Linki tab

    def _build_linki_tab(self) -> QWidget:
        from ui.dict_tab import DictTab

        def excel_parser(ws):
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            result = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                for col_idx, fleet in enumerate(headers):
                    if not fleet:
                        continue
                    url = str(row[col_idx] or "").strip() if col_idx < len(row) else ""
                    if url and url.lower() not in ("none", ""):
                        result.append([fleet, url])
            return result

        return DictTab(
            headers=["Flota", "Link"],
            loader=self._db.get_all_fleet_links,
            saver=lambda vals: self._db.upsert_fleet_link(vals[0], vals[1] if len(vals) > 1 else ""),
            deleter=self._db.delete_fleet_link_by_id,
            excel_sheet="Linki",
            excel_parser=excel_parser,
            add_labels=["Flota", "Link"],
            commit_fn=self._db.commit,
        )

    # ---------------------------------------------------------------- Zbiorczy import tab

    def _get_bulk_configs(self) -> list[dict]:
        from database.models import Technician

        def parse_firmy(ws):
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            result = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                for col_idx, fleet in enumerate(headers):
                    if not fleet:
                        continue
                    company = str(row[col_idx] or "").strip() if col_idx < len(row) else ""
                    if company and company.lower() not in ("none", ""):
                        result.append([company, fleet])
            return result

        def parse_pojazdy(ws):
            result = []
            for row in ws.iter_rows(values_only=True):
                model = str(row[0] or "").strip()
                vtype = str(row[1] or "").strip() if len(row) > 1 else ""
                if model and model.lower() not in ("none", "marka", "model"):
                    result.append([model, vtype])
            return result

        def parse_single(*skip_words):
            skip = set(w.lower() for w in skip_words) | {"none"}
            def _parser(ws):
                result = []
                for row in ws.iter_rows(values_only=True):
                    val = str(row[0] or "").strip()
                    if val and val.lower() not in skip:
                        result.append([val])
                return result
            return _parser

        def parse_extra_devices(ws):
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            result = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                for col_idx, fleet in enumerate(headers):
                    if not fleet:
                        continue
                    device = str(row[col_idx] or "").strip() if col_idx < len(row) else ""
                    if device and device.lower() not in ("none", ""):
                        result.append([device, fleet])
            return result

        def parse_linki(ws):
            headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]
            result = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                for col_idx, fleet in enumerate(headers):
                    if not fleet:
                        continue
                    url = str(row[col_idx] or "").strip() if col_idx < len(row) else ""
                    if url and url.lower() not in ("none", ""):
                        result.append([fleet, url])
            return result

        db = self._db
        return [
            {
                "name": "Firmy",
                "sheet": "Firmy",
                "parser": parse_firmy,
                "saver": lambda v: db.upsert_company_with_fleet(v[0], v[1] if len(v) > 1 else ""),
            },
            {
                "name": "Pojazdy",
                "sheet": "Model_Typ",
                "parser": parse_pojazdy,
                "saver": lambda v: db.upsert_vehicle_model(v[0], v[1] if len(v) > 1 else ""),
            },
            {
                "name": "Monterzy",
                "sheet": "Monterzy",
                "parser": parse_single("monter"),
                "saver": lambda v: db.upsert_technician(Technician(full_name=v[0])),
            },
            {
                "name": "Gdzie rejestrator",
                "sheet": "GdzieRejestrator",
                "parser": parse_single("lokalizacja"),
                "saver": lambda v: db.upsert_recorder_location(v[0]),
            },
            {
                "name": "Model urządzenia",
                "sheet": "ModelUrzadzenia",
                "parser": parse_single("model", "model urządzenia"),
                "saver": lambda v: db.upsert_device_model(v[0]),
            },
            {
                "name": "Urządzenia dodatkowe",
                "sheet": "UrzadzeniaDodatkowe",
                "parser": parse_extra_devices,
                "saver": lambda v: db.upsert_extra_device(v[1] if len(v) > 1 else "", v[0]),
            },
            {
                "name": "Linki flot",
                "sheet": "Linki",
                "parser": parse_linki,
                "saver": lambda v: db.upsert_fleet_link(v[0], v[1] if len(v) > 1 else ""),
            },
        ]

    def _build_zbiorcze_tab(self) -> QWidget:
        w = QWidget()
        outer = QVBoxLayout(w)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # === Zakładki słowników na górze, wypełniają całe okno ===
        self._dict_tab_refs: list = []
        sub = QTabWidget()
        sub.setDocumentMode(True)

        for label, build_fn in [
            ("Firmy",                self._build_firmy_tab),
            ("Pojazdy",              self._build_pojazdy_tab),
            ("Monterzy",             self._build_monterzy_tab),
            ("Gdzie rejestrator",    self._build_loki_tab),
            ("Model urządzenia",     self._build_device_models_tab),
            ("Urządzenia dodatkowe", self._build_extra_devices_tab),
            ("Linki flot",           self._build_linki_tab),
        ]:
            tab = build_fn()
            self._dict_tab_refs.append(tab)
            sub.addTab(tab, f"  {label}  ")

        outer.addWidget(sub, 1)
        return w

    def _build_bulk_import_widget(self) -> QWidget:
        from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView

        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(14, 14, 14, 14)
        lay.setSpacing(10)

        hint = QLabel(
            "Wybierz plik Excel zawierający arkusze słownikowe.\n"
            "Wszystkie pasujące arkusze zostaną zaimportowane jednocześnie — brakujące są pomijane."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #94a3b8; font-size: 8.5pt;")
        lay.addWidget(hint)

        file_row = QHBoxLayout()
        self._bulk_path = QLineEdit()
        self._bulk_path.setPlaceholderText("Ścieżka do pliku .xlsx / .xlsm …")
        file_row.addWidget(self._bulk_path, 1)
        btn_browse = QPushButton("Przeglądaj…")
        btn_browse.setFixedWidth(90)
        btn_browse.clicked.connect(self._on_bulk_browse)
        file_row.addWidget(btn_browse)
        lay.addLayout(file_row)

        btn_row = QHBoxLayout()
        btn_import = QPushButton("📥  Importuj wszystkie arkusze")
        btn_import.setObjectName("btn_primary")
        btn_import.clicked.connect(self._on_bulk_import)
        btn_refresh = QPushButton("🔄  Odśwież słowniki")
        btn_refresh.clicked.connect(self._on_refresh_all)
        btn_row.addWidget(btn_import, 1)
        btn_row.addWidget(btn_refresh)
        lay.addLayout(btn_row)

        self._bulk_table = QTableWidget(0, 3)
        self._bulk_table.setHorizontalHeaderLabels(["Słownik", "Arkusz w pliku", "Wynik"])
        self._bulk_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._bulk_table.setSelectionMode(QAbstractItemView.NoSelection)
        self._bulk_table.verticalHeader().setVisible(False)
        self._bulk_table.verticalHeader().setDefaultSectionSize(24)
        hdr = self._bulk_table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)

        configs = self._get_bulk_configs()
        self._bulk_table.setRowCount(len(configs))
        for i, cfg in enumerate(configs):
            self._bulk_table.setItem(i, 0, QTableWidgetItem(cfg["name"]))
            self._bulk_table.setItem(i, 1, QTableWidgetItem(cfg["sheet"]))
            self._bulk_table.setItem(i, 2, QTableWidgetItem("–"))

        lay.addWidget(self._bulk_table, 1)
        return w

    def _on_refresh_all(self):
        for tab in getattr(self, "_dict_tab_refs", []):
            if hasattr(tab, "refresh"):
                tab.refresh()

    def _on_bulk_browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik Excel", "",
            "Pliki Excel (*.xlsx *.xlsm *.xls)"
        )
        if path:
            self._bulk_path.setText(path)

    def _on_bulk_import(self):
        import shutil
        import tempfile

        path = self._bulk_path.text().strip()
        if not path:
            path, _ = QFileDialog.getOpenFileName(
                self, "Wybierz plik Excel", "",
                "Pliki Excel (*.xlsx *.xlsm *.xls)"
            )
            if not path:
                return
            self._bulk_path.setText(path)

        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "Brak biblioteki", "Uruchom: pip install openpyxl")
            return

        tmp_path = None
        try:
            suffix = os.path.splitext(path)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_f:
                tmp_path = tmp_f.name
            shutil.copy2(path, tmp_path)
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        except Exception as e:
            QMessageBox.critical(self, "Błąd odczytu", str(e))
            return

        configs = self._get_bulk_configs()
        total = 0

        for i, cfg in enumerate(configs):
            if cfg["sheet"] not in wb.sheetnames:
                self._bulk_table.item(i, 2).setText("brak arkusza")
                continue
            try:
                rows = cfg["parser"](wb[cfg["sheet"]])
                count = sum(
                    1 for vals in rows
                    if any(v for v in vals if v) and (cfg["saver"](vals) or True)
                )
                total += count
                self._bulk_table.item(i, 2).setText(f"✓  {count} rekordów")
            except Exception as e:
                self._bulk_table.item(i, 2).setText(f"Błąd: {e}")
                logger.error(f"Bulk import error [{cfg['name']}]: {e}", exc_info=True)

        wb.close()
        if tmp_path:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

        self._db.commit()
        self._on_refresh_all()
        QMessageBox.information(
            self, "Import zakończony",
            f"Zaimportowano łącznie {total} rekordów."
        )

    # ---------------------------------------------------------------- Import Odbiory tab

    def _build_odbiory_import_tab(self) -> QWidget:
        from PySide6.QtWidgets import QProgressBar
        from PySide6.QtCore import QSettings as _QS

        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(14, 14, 14, 14)
        lay.setSpacing(10)

        hint = QLabel(
            "Import rekordów z pliku Odbiory.xlsb do bazy danych.\n"
            "Każdy wiersz zostaje zaimportowany — bez pomijania duplikatów.\n"
            "Kolumny dopasowywane są automatycznie po nagłówkach (obsługa różnych układów pliku)."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #94a3b8; font-size: 8.5pt;")
        lay.addWidget(hint)

        file_row = QHBoxLayout()
        self._import_path = QLineEdit()
        self._import_path.setPlaceholderText("Ścieżka do pliku Odbiory.xlsb …")
        saved_path = _QS("TwojaFirma", "SystemOdbiory").value("import/odbiory_path", "")
        self._import_path.setText(saved_path)
        file_row.addWidget(self._import_path, 1)

        btn_browse = QPushButton("Przeglądaj…")
        btn_browse.setFixedWidth(90)
        btn_browse.clicked.connect(self._on_import_browse)
        file_row.addWidget(btn_browse)
        lay.addLayout(file_row)

        opt_row = QHBoxLayout()
        self._btn_import_odbiory = QPushButton("📥  Importuj Odbiory.xlsb")
        self._btn_import_odbiory.setObjectName("btn_primary")
        self._btn_import_odbiory.clicked.connect(self._on_import_odbiory)
        opt_row.addWidget(self._btn_import_odbiory, 1)
        lay.addLayout(opt_row)

        self._import_progress_bar = QProgressBar()
        self._import_progress_bar.setVisible(False)
        self._import_progress_bar.setFixedHeight(14)
        lay.addWidget(self._import_progress_bar)

        self._import_progress_lbl = QLabel("")
        self._import_progress_lbl.setStyleSheet("color: #64748b; font-size: 8.5pt;")
        lay.addWidget(self._import_progress_lbl)

        self._import_result_lbl = QLabel("")
        self._import_result_lbl.setWordWrap(True)
        self._import_result_lbl.setStyleSheet("font-size: 9.5pt;")
        lay.addWidget(self._import_result_lbl)

        lay.addStretch()

        note = QLabel("Wymagana biblioteka: pip install pyxlsb")
        note.setStyleSheet("color: #64748b; font-size: 8pt;")
        lay.addWidget(note)

        return w

    def _on_import_browse(self):
        from PySide6.QtCore import QSettings as _QS
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik Odbiory.xlsb", "",
            "Pliki Excel Binary (*.xlsb)"
        )
        if path:
            self._import_path.setText(path)
            _QS("TwojaFirma", "SystemOdbiory").setValue("import/odbiory_path", path)

    def _on_import_odbiory(self):
        from PySide6.QtCore import QSettings as _QS
        path = self._import_path.text().strip()
        if not path:
            path, _ = QFileDialog.getOpenFileName(
                self, "Wybierz plik Odbiory.xlsb", "",
                "Pliki Excel Binary (*.xlsb)"
            )
            if not path:
                return
            self._import_path.setText(path)
            _QS("TwojaFirma", "SystemOdbiory").setValue("import/odbiory_path", path)

        if not os.path.isfile(path):
            QMessageBox.warning(self, "Plik nie istnieje", f"Nie znaleziono pliku:\n{path}")
            return

        if self._import_worker and self._import_worker.isRunning():
            return

        self._btn_import_odbiory.setEnabled(False)
        self._import_progress_lbl.setText("Uruchamianie importu…")
        self._import_progress_bar.setRange(0, 0)
        self._import_progress_bar.setVisible(True)
        self._import_result_lbl.setText("")

        self._import_worker = OdbioryImportWorker(path)
        self._import_worker.progress.connect(self._import_progress_lbl.setText)
        self._import_worker.row_done.connect(self._on_import_row_done)
        self._import_worker.finished.connect(self._on_import_finished)
        self._import_worker.start()

    @Slot(int, int)
    def _on_import_row_done(self, current: int, total: int):
        if self._import_progress_bar.maximum() == 0:
            self._import_progress_bar.setRange(0, total)
        self._import_progress_bar.setValue(current)
        if current % 100 == 0 or current == total:
            self._import_progress_lbl.setText(f"Przetwarzanie: {current} / {total}…")

    @Slot(bool, str, int)
    def _on_import_finished(self, ok: bool, message: str, errors: int):
        self._btn_import_odbiory.setEnabled(True)
        self._import_progress_bar.setVisible(False)
        self._import_progress_lbl.setText("")

        color = "#22c55e" if (ok and errors == 0) else "#f59e0b" if ok else "#ef4444"
        self._import_result_lbl.setStyleSheet(f"font-size: 9.5pt; color: {color};")
        self._import_result_lbl.setText(message)

    # ---------------------------------------------------------------- Columns tab

    def _build_columns_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(14, 14, 14, 10)
        lay.setSpacing(8)

        hint = QLabel("Zaznacz widoczne kolumny · przeciągaj lub użyj przycisków ▲▼ aby zmienić kolejność")
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #94a3b8; font-size: 8.5pt;")
        lay.addWidget(hint)

        self._col_list = QListWidget()
        self._col_list.setDragDropMode(QAbstractItemView.InternalMove)
        self._col_list.setDefaultDropAction(Qt.MoveAction)
        self._col_list.setSelectionMode(QAbstractItemView.SingleSelection)
        self._col_list.setStyleSheet(
            "QListWidget { font-size: 9pt; }"
            "QListWidget::item { padding: 4px 6px; }"
        )

        label_map = {attr: lbl for lbl, attr, _ in self._all_columns}
        for attr in self._column_order:
            if attr == "id":
                continue
            item = QListWidgetItem(label_map.get(attr, attr))
            item.setData(Qt.UserRole, attr)
            item.setFlags(
                Qt.ItemIsEnabled | Qt.ItemIsSelectable
                | Qt.ItemIsUserCheckable | Qt.ItemIsDragEnabled
            )
            item.setCheckState(
                Qt.Checked if attr in self._visible_set else Qt.Unchecked
            )
            self._col_list.addItem(item)

        lay.addWidget(self._col_list, 1)

        move_row = QHBoxLayout()
        btn_up = QPushButton("▲  Wyżej")
        btn_down = QPushButton("▼  Niżej")
        btn_up.clicked.connect(self._move_col_up)
        btn_down.clicked.connect(self._move_col_down)
        move_row.addWidget(btn_up)
        move_row.addWidget(btn_down)
        move_row.addStretch()

        btn_all = QPushButton("Zaznacz wszystkie")
        btn_none = QPushButton("Odznacz wszystkie")
        btn_all.clicked.connect(lambda: self._set_all_checked(True))
        btn_none.clicked.connect(lambda: self._set_all_checked(False))
        move_row.addWidget(btn_all)
        move_row.addWidget(btn_none)
        lay.addLayout(move_row)

        btn_apply = QPushButton("Zastosuj kolumny")
        btn_apply.setObjectName("btn_primary")
        btn_apply.clicked.connect(self._apply_columns)
        lay.addWidget(btn_apply)

        return w

    # ---------------------------------------------------------------- General settings tab

    # ---------------------------------------------------------------- Colors tab

    _COLOR_DEFAULTS = {
        "color_odebrane_light": "#dcfce7",
        "color_odebrane_dark":  "#1e6640",
        "odebrane_highlight_mode": "row",          # "row" | "cols" | "none"
        "odebrane_highlight_cols": "company_name",
        "dyzur_highlight_enabled": "1",
        "dyzur_highlight_cols": "_datetime",
        "color_dyzur_light": "#fef3c7",
        "color_dyzur_dark":  "#78350f",
    }

    def _build_colors_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(14, 14, 14, 10)
        lay.setSpacing(14)

        # ── Grupa 1: Odebrane ────────────────────────────────────────────────
        grp_od = QGroupBox("Podświetlenie wiersza 'Odebrane'")
        gl_od = QGridLayout(grp_od)
        gl_od.setContentsMargins(12, 16, 12, 12)
        gl_od.setHorizontalSpacing(10)
        gl_od.setVerticalSpacing(10)
        gl_od.setColumnStretch(2, 1)

        gl_od.addWidget(QLabel("Tryb:"), 0, 0)
        mode_w = QWidget(); mode_w.setStyleSheet("background:transparent;")
        mode_lay = QHBoxLayout(mode_w); mode_lay.setContentsMargins(0,0,0,0); mode_lay.setSpacing(10)
        self._rb_od_row   = QRadioButton("Cały wiersz")
        self._rb_od_cols  = QRadioButton("Wybrane kolumny:")
        self._rb_od_none  = QRadioButton("Brak")
        self._od_mode_grp = QButtonGroup(self)
        for rb in (self._rb_od_row, self._rb_od_cols, self._rb_od_none):
            self._od_mode_grp.addButton(rb)

        mode_lay.addWidget(self._rb_od_row)
        mode_lay.addWidget(self._rb_od_cols)
        
        self._lbl_od_cols = QLabel()
        self._lbl_od_cols.setStyleSheet("color: #64748b; font-size: 8.5pt;")
        mode_lay.addWidget(self._lbl_od_cols)
        
        self._btn_od_cols_pick = QPushButton("Zmień...")
        self._btn_od_cols_pick.setStyleSheet("font-size: 8pt;")
        self._btn_od_cols_pick.clicked.connect(lambda: self._pick_highlight_columns("odebrane_highlight_cols", self._lbl_od_cols))
        mode_lay.addWidget(self._btn_od_cols_pick)
        
        mode_lay.addWidget(self._rb_od_none)
        mode_lay.addStretch()
        current_mode = self._db.get_setting("odebrane_highlight_mode", "row")
        if current_mode == "firma": current_mode = "cols"
        {"row": self._rb_od_row, "cols": self._rb_od_cols, "none": self._rb_od_none}.get(
            current_mode, self._rb_od_row
        ).setChecked(True)
        self._od_mode_grp.buttonClicked.connect(self._on_od_mode_changed)
        gl_od.addWidget(mode_w, 0, 1, 1, 3)

        self._update_cols_label("odebrane_highlight_cols", self._lbl_od_cols)

        self._clr_od_light_prev, self._clr_od_light_lbl = self._add_color_row(
            gl_od, 1, "Kolor (jasny):", "color_odebrane_light", self._COLOR_DEFAULTS["color_odebrane_light"]
        )
        self._clr_od_dark_prev, self._clr_od_dark_lbl = self._add_color_row(
            gl_od, 2, "Kolor (ciemny):", "color_odebrane_dark", self._COLOR_DEFAULTS["color_odebrane_dark"]
        )

        lay.addWidget(grp_od)

        # ── Grupa 2: Dyżur ───────────────────────────────────────────────────
        grp_dz = QGroupBox("Podświetlenie wierszy dla 'Dyżur'")
        gl_dz = QGridLayout(grp_dz)
        gl_dz.setContentsMargins(12, 16, 12, 12)
        gl_dz.setHorizontalSpacing(10)
        gl_dz.setVerticalSpacing(10)
        gl_dz.setColumnStretch(2, 1)

        self._cb_dyzur_color = QCheckBox("Włącz podświetlanie dla dyżurów")
        self._cb_dyzur_color.setChecked(
            self._db.get_setting("dyzur_highlight_enabled", "1") == "1"
        )
        self._cb_dyzur_color.toggled.connect(
            lambda c: self._db.set_setting("dyzur_highlight_enabled", "1" if c else "0")
        )
        gl_dz.addWidget(self._cb_dyzur_color, 0, 0, 1, 4)

        dz_cols_w = QWidget(); dz_cols_w.setStyleSheet("background:transparent;")
        dz_cols_lay = QHBoxLayout(dz_cols_w); dz_cols_lay.setContentsMargins(24,0,0,0); dz_cols_lay.setSpacing(10)
        dz_cols_lay.addWidget(QLabel("Wybrane kolumny:"))
        self._lbl_dz_cols = QLabel()
        self._lbl_dz_cols.setStyleSheet("color: #64748b; font-size: 8.5pt;")
        dz_cols_lay.addWidget(self._lbl_dz_cols)

        self._btn_dz_cols_pick = QPushButton("Zmień...")
        self._btn_dz_cols_pick.setStyleSheet("font-size: 8pt;")
        self._btn_dz_cols_pick.clicked.connect(lambda: self._pick_highlight_columns("dyzur_highlight_cols", self._lbl_dz_cols))
        dz_cols_lay.addWidget(self._btn_dz_cols_pick)
        dz_cols_lay.addStretch()
        gl_dz.addWidget(dz_cols_w, 1, 0, 1, 4)
        self._update_cols_label("dyzur_highlight_cols", self._lbl_dz_cols)

        self._clr_dz_light_prev, self._clr_dz_light_lbl = self._add_color_row(
            gl_dz, 2, "Kolor (jasny):", "color_dyzur_light", self._COLOR_DEFAULTS["color_dyzur_light"]
        )
        self._clr_dz_dark_prev, self._clr_dz_dark_lbl = self._add_color_row(
            gl_dz, 3, "Kolor (ciemny):", "color_dyzur_dark", self._COLOR_DEFAULTS["color_dyzur_dark"]
        )

        lay.addWidget(grp_dz)

        # ── Reset ────────────────────────────────────────────────────────────
        btn_reset = QPushButton("Przywróć wszystkie domyślne")
        btn_reset.setFixedWidth(190)
        btn_reset.clicked.connect(self._reset_colors)
        lay.addWidget(btn_reset, alignment=Qt.AlignLeft)

        lay.addStretch()
        return w

    def _add_color_row(self, grid: QGridLayout, row: int, label: str,
                       key: str, default: str):
        """Dodaje wiersz z podglądem koloru i przyciskiem Zmień do siatki. Zwraca (preview, lbl)."""
        current = self._db.get_setting(key, default)
        grid.addWidget(QLabel(label), row, 0)

        preview = QPushButton()
        preview.setFixedSize(34, 22)
        preview.setFlat(True)
        preview.setStyleSheet(
            f"background:{current}; border:1px solid #64748b; border-radius:3px;"
        )
        grid.addWidget(preview, row, 1)

        lbl = QLabel(current)
        lbl.setStyleSheet("font-family: monospace; font-size: 9pt;")
        grid.addWidget(lbl, row, 2)

        btn = QPushButton("Zmień…")
        btn.setStyleSheet("font-size: 8pt;")
        btn.clicked.connect(lambda _=False, k=key, p=preview, l=lbl: self._pick_color(k, p, l))
        grid.addWidget(btn, row, 3)

        return preview, lbl

    def _on_od_mode_changed(self):
        if self._rb_od_row.isChecked():     mode = "row"
        elif self._rb_od_cols.isChecked():  mode = "cols"
        else:                               mode = "none"
        self._db.set_setting("odebrane_highlight_mode", mode)

    def _update_cols_label(self, key: str, lbl: QLabel):
        val = self._db.get_setting(key, self._COLOR_DEFAULTS.get(key, ""))
        attrs = [x.strip() for x in val.split(",") if x.strip()]
        attr_to_lbl = {a: l for l, a, _ in self._all_columns}
        labels = [attr_to_lbl.get(a, a) for a in attrs]
        lbl.setText(", ".join(labels) if labels else "Brak")

    def _pick_highlight_columns(self, key: str, lbl: QLabel):
        dlg = QDialog(self)
        dlg.setWindowTitle("Wybierz kolumny do podświetlenia")
        dlg.setModal(True)
        dlg.resize(300, 450)
        lay = QVBoxLayout(dlg)

        list_w = QListWidget()
        list_w.setSelectionMode(QAbstractItemView.NoSelection)

        val = self._db.get_setting(key, self._COLOR_DEFAULTS.get(key, ""))
        selected_attrs = set(x.strip() for x in val.split(",") if x.strip())

        for label, attr, _ in self._all_columns:
            if attr == "id": continue
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, attr)
            item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if attr in selected_attrs else Qt.Unchecked)
            list_w.addItem(item)

        lay.addWidget(list_w)

        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(dlg.accept)
        btn_box.rejected.connect(dlg.reject)
        lay.addWidget(btn_box)

        if dlg.exec() == QDialog.Accepted:
            new_attrs = []
            for i in range(list_w.count()):
                item = list_w.item(i)
                if item.checkState() == Qt.Checked:
                    new_attrs.append(item.data(Qt.UserRole))
            self._db.set_setting(key, ",".join(new_attrs))
            self._update_cols_label(key, lbl)
            self._on_od_mode_changed()

    def _pick_color(self, key: str, preview: QPushButton, lbl: QLabel):
        color = QColorDialog.getColor(QColor(lbl.text()), self, "Wybierz kolor")
        if not color.isValid():
            return
        hex_color = color.name()
        self._db.set_setting(key, hex_color)
        preview.setStyleSheet(f"background:{hex_color}; border:1px solid #64748b; border-radius:3px;")
        lbl.setText(hex_color)

    def _reset_colors(self):
        color_map = {
            "color_odebrane_light": (self._clr_od_light_prev, self._clr_od_light_lbl),
            "color_odebrane_dark":  (self._clr_od_dark_prev,  self._clr_od_dark_lbl),
            "color_dyzur_light":    (self._clr_dz_light_prev, self._clr_dz_light_lbl),
            "color_dyzur_dark":     (self._clr_dz_dark_prev,  self._clr_dz_dark_lbl),
        }
        for key, (preview, lbl) in color_map.items():
            default = self._COLOR_DEFAULTS[key]
            self._db.set_setting(key, default)
            preview.setStyleSheet(f"background:{default}; border:1px solid #64748b; border-radius:3px;")
            lbl.setText(default)
        self._db.set_setting("odebrane_highlight_mode", "row")
        self._db.set_setting("odebrane_highlight_cols", self._COLOR_DEFAULTS["odebrane_highlight_cols"])
        self._rb_od_row.setChecked(True)
        self._update_cols_label("odebrane_highlight_cols", self._lbl_od_cols)
        
        self._db.set_setting("dyzur_highlight_enabled", "1")
        self._db.set_setting("dyzur_highlight_cols", self._COLOR_DEFAULTS["dyzur_highlight_cols"])
        self._cb_dyzur_color.setChecked(True)
        self._update_cols_label("dyzur_highlight_cols", self._lbl_dz_cols)

    # ---------------------------------------------------------------- General tab

    def _build_general_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(14, 14, 14, 10)
        lay.setSpacing(16)

        # Bardziej widoczne, rozróżnialne checkboxy
        cb_style = f"""
            QCheckBox {{
                font-size: 10pt;
                padding: 4px;
                color: {'#0f172a' if self._is_light else '#e2e8f0'};
            }}
            QCheckBox::indicator {{
                width: 22px;
                height: 22px;
                border-radius: 4px;
            border: 2px solid {'#cbd5e1' if self._is_light else '#64748b'};
                background-color: {'#ffffff' if self._is_light else '#1e2229'};
            }}
            QCheckBox::indicator:hover {{
                border-color: {'#94a3b8' if self._is_light else '#94a3b8'};
            }}
            QCheckBox::indicator:checked {{
            background-color: {'#ffffff' if self._is_light else '#64748b'};
            border-color: {'#cbd5e1' if self._is_light else '#94a3b8'};
            image: url("data:image/svg+xml;utf8,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 24 24' fill='none' stroke='{'%230f172a' if self._is_light else 'white'}' stroke-width='4' stroke-linecap='round' stroke-linejoin='round'><polyline points='20 6 9 17 4 12'/></svg>");
            }}
        """

        self._cb_duty = QCheckBox(" Dyżurny")
        self._cb_duty.setStyleSheet(cb_style)
        
        # Pobranie ustawienia z bazy (domyślnie włączone - "1")
        is_duty_enabled = self._db.get_setting("show_duty_section", "1") == "1"
        self._cb_duty.setChecked(is_duty_enabled)
        self._cb_duty.toggled.connect(self._on_duty_toggled)

        lay.addWidget(self._cb_duty)

        self._cb_theme = QCheckBox(" Jasny motyw aplikacji (Bright Mode)")
        self._cb_theme.setStyleSheet(cb_style)
        self._cb_theme.setChecked(self._is_light)
        self._cb_theme.toggled.connect(self._on_theme_toggled)
        lay.addWidget(self._cb_theme)

        info = QLabel("ℹ  Zmiana motywu wymaga ponownego uruchomienia aplikacji (lub otwarcia nowych okien), aby w pełni zadziałała.")
        info.setStyleSheet("color: #64748b; font-size: 9pt; margin-top: 10px;")
        info.setWordWrap(True)
        lay.addWidget(info)

        # --- Filtry ---
        grp_filters = QGroupBox("Filtry")
        fl = QVBoxLayout(grp_filters)
        fl.setSpacing(6)

        self._cb_remember_search = QCheckBox(" Pamiętaj filtr wyszukiwarki między sesjami")
        self._cb_remember_search.setStyleSheet(cb_style)
        self._cb_remember_search.setChecked(self._db.get_setting("remember_search_filter", "1") == "1")
        self._cb_remember_search.toggled.connect(
            lambda v: self._db.set_setting("remember_search_filter", "1" if v else "0")
        )
        fl.addWidget(self._cb_remember_search)

        self._cb_remember_dates = QCheckBox(" Pamiętaj filtry dat między sesjami")
        self._cb_remember_dates.setStyleSheet(cb_style)
        self._cb_remember_dates.setChecked(self._db.get_setting("remember_date_filter", "1") == "1")
        self._cb_remember_dates.toggled.connect(
            lambda v: self._db.set_setting("remember_date_filter", "1" if v else "0")
        )
        fl.addWidget(self._cb_remember_dates)

        lay.addWidget(grp_filters)

        # --- Kopia zapasowa ---
        grp_backup = QGroupBox("Kopia zapasowa bazy danych")
        bl = QVBoxLayout(grp_backup)
        bl.setSpacing(8)

        btn_row_b = QHBoxLayout()
        btn_backup = QPushButton("💾  Utwórz kopię zapasową")
        btn_backup.clicked.connect(self._on_create_backup)

        btn_restore = QPushButton("📂  Przywróć z kopii")
        btn_restore.clicked.connect(self._on_restore_backup)

        btn_row_b.addWidget(btn_backup)
        btn_row_b.addWidget(btn_restore)
        btn_row_b.addStretch()
        bl.addLayout(btn_row_b)

        auto_row = QHBoxLayout()
        auto_row.addWidget(QLabel("Automatyczna kopia (folder):"))
        self._auto_backup_edit = QLineEdit()
        self._auto_backup_edit.setPlaceholderText("Wybierz folder…")
        self._auto_backup_edit.setText(self._db.get_setting("auto_backup_path", ""))
        self._auto_backup_edit.textChanged.connect(
            lambda t: self._db.set_setting("auto_backup_path", t.strip())
        )
        auto_row.addWidget(self._auto_backup_edit, 1)
        btn_auto_browse = QPushButton("…")
        btn_auto_browse.setFixedWidth(28)
        btn_auto_browse.clicked.connect(self._on_auto_backup_browse)
        auto_row.addWidget(btn_auto_browse)
        bl.addLayout(auto_row)

        note_auto = QLabel("ℹ  Kopia tworzona automatycznie przy każdym zamknięciu aplikacji.")
        note_auto.setStyleSheet("color: #64748b; font-size: 8pt;")
        bl.addWidget(note_auto)

        lay.addWidget(grp_backup)

        # --- Aktualizacje ---
        grp_update = QGroupBox("Aktualizacje")
        ul = QVBoxLayout(grp_update)
        ul.setSpacing(8)

        btn_update = QPushButton("🔄  Sprawdź aktualizacje")
        btn_update.clicked.connect(self._on_check_updates)
        ul.addWidget(btn_update)
        lay.addWidget(grp_update)

        lay.addStretch()

        return w

    # ---------------------------------------------------------------- O aplikacji tab

    def _build_about_tab(self) -> QWidget:
        from PySide6.QtWidgets import QScrollArea
        from changelog import CHANGELOG
        from ui.whats_new_dialog import _make_collapsible_section

        w = QWidget()
        outer = QVBoxLayout(w)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)

        content = QWidget()
        lay = QVBoxLayout(content)
        lay.setContentsMargins(20, 16, 20, 16)
        lay.setSpacing(0)

        for i, entry in enumerate(CHANGELOG[:5]):
            if i > 0:
                sep = QWidget()
                sep.setFixedHeight(1)
                sep.setStyleSheet("background: #2e3340; margin: 4px 0;")
                lay.addWidget(sep)
            _make_collapsible_section(lay, entry, expanded=(i == 0))

        lay.addStretch()
        scroll.setWidget(content)
        outer.addWidget(scroll, 1)
        return w

    # ---------------------------------------------------------------- helpers – Auto backup

    def _on_auto_backup_browse(self):
        from PySide6.QtWidgets import QFileDialog as _QFD
        folder = _QFD.getExistingDirectory(self, "Wybierz folder automatycznej kopii zapasowej")
        if folder:
            self._auto_backup_edit.setText(folder)

    # ---------------------------------------------------------------- helpers – SIM

    def _on_browse(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik Excel", "",
            "Pliki Excel (*.xlsx *.xlsm *.xls)"
        )
        if path:
            self._file_edit.setText(path)

    def _on_sync_url(self):
        url = self._url_edit.text().strip()
        if not url:
            self._progress_lbl.setText("Podaj URL pliku.")
            return
        self._db.set_setting("sim_source_url", url)
        self._start_sync(url, is_url=True)

    def _on_sync_file(self):
        path = self._file_edit.text().strip()
        if not path or not os.path.isfile(path):
            self._progress_lbl.setText("Wskaż istniejący plik Excel.")
            return
        self._db.set_setting("sim_source_file", path)
        self._start_sync(path, is_url=False)

    def _start_sync(self, source: str, is_url: bool):
        if self._worker and self._worker.isRunning():
            return
        self._progress_lbl.setText("Synchronizacja w toku…")
        self._worker = SyncWorker(source, is_url)
        self._worker.progress.connect(self._progress_lbl.setText)
        self._worker.finished.connect(self._on_sync_done)
        self._worker.start()

    @Slot(bool, str)
    def _on_sync_done(self, ok: bool, message: str):
        self._progress_lbl.setText(message)
        if ok:
            self._lbl_last.setText(
                f"Ostatnia synchronizacja:  {self._db.get_setting('sim_last_sync', '–')}"
            )
            self._lbl_count.setText(f"Kart SIM w bazie:  {self._db.get_sim_cards_count()}")
            self._sim_dict_tab.refresh()

    def _on_clear_sim(self):
        reply = QMessageBox.question(
            self, "Wyczyść bazę SIM",
            "Czy na pewno chcesz usunąć wszystkie karty SIM z bazy?\nTej operacji nie można cofnąć.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            self._db.clear_sim_cards()
            self._lbl_count.setText("Kart SIM w bazie:  0")
            self._progress_lbl.setText("Baza SIM wyczyszczona.")
            self._sim_dict_tab.refresh()

    def _on_create_backup(self):
        from config import DB_PATH
        import shutil
        from datetime import datetime
        default_name = f"odbiory_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        path, _ = QFileDialog.getSaveFileName(
            self, "Zapisz kopię zapasową", default_name, "Baza danych SQLite (*.db *.sqlite)"
        )
        if path:
            try:
                shutil.copy2(DB_PATH, path)
                QMessageBox.information(self, "Kopia zapasowa", f"Pomyślnie utworzono kopię zapasową:\n{path}")
            except Exception as e:
                QMessageBox.critical(self, "Błąd", f"Nie udało się utworzyć kopii:\n{e}")

    def _on_restore_backup(self):
        from config import DB_PATH
        import shutil
        from PySide6.QtWidgets import QApplication
        path, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik kopii zapasowej", "", "Baza danych SQLite (*.db *.sqlite);;Wszystkie pliki (*.*)"
        )
        if path:
            if QMessageBox.warning(
                self, "Przywracanie bazy danych",
                "UWAGA: Aktualna baza danych zostanie całkowicie nadpisana!\n\nOperacja wymaga ponownego uruchomienia aplikacji.\nCzy na pewno chcesz kontynuować?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            ) == QMessageBox.Yes:
                try:
                    self._db.close() # Zamknięcie aktywnego połączenia przed nadpisaniem
                    shutil.copy2(path, DB_PATH)
                    QMessageBox.information(self, "Sukces", "Baza danych została przywrócona.\nAplikacja zostanie teraz zamknięta.")
                    QApplication.quit()
                except Exception as e:
                    QMessageBox.critical(self, "Błąd", f"Nie udało się przywrócić bazy:\n{e}")

    def _on_check_updates(self):
        from config import APP_VERSION, VERSION_URL, INSTALLER_URL
        try:
            import requests
        except ImportError:
            QMessageBox.warning(self, "Brak biblioteki",
                                "Brakuje biblioteki 'requests'.\nUruchom: pip install requests")
            return

        try:
            r = requests.get(VERSION_URL, timeout=8)
            r.raise_for_status()
            latest = r.text.strip()
        except Exception as e:
            QMessageBox.warning(self, "Błąd połączenia",
                                f"Nie udało się sprawdzić aktualizacji:\n{e}")
            return

        def _ver_tuple(v: str):
            try:
                return tuple(int(x) for x in v.strip().split("."))
            except ValueError:
                return (0,)

        if _ver_tuple(latest) <= _ver_tuple(APP_VERSION):
            QMessageBox.information(self, "Aktualizacja",
                                    f"Posiadasz najnowszą wersję ({APP_VERSION}).")
            return

        reply = QMessageBox.question(
            self, "Dostępna nowa wersja",
            f"Dostępna wersja:  {latest}\n"
            f"Twoja wersja:      {APP_VERSION}\n\n"
            f"Pobrać i zainstalować automatycznie?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes,
        )
        if reply == QMessageBox.Yes:
            self._download_and_install(INSTALLER_URL, latest)

    def _download_and_install(self, url: str, version: str):
        import os, tempfile, subprocess
        from PySide6.QtCore import Qt

        tmp_path = os.path.join(tempfile.gettempdir(), "Odbiory_Setup.exe")

        progress = QProgressDialog(
            f"Pobieranie wersji {version}…", "Anuluj", 0, 100, self
        )
        progress.setWindowTitle("Aktualizacja")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumWidth(340)
        progress.setValue(0)
        progress.show()

        try:
            import requests
            r = requests.get(url, stream=True, timeout=60)
            r.raise_for_status()
            total = int(r.headers.get("content-length", 0))
            downloaded = 0

            with open(tmp_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=65536):
                    if progress.wasCanceled():
                        return
                    f.write(chunk)
                    downloaded += len(chunk)
                    if total:
                        progress.setValue(int(downloaded * 100 / total))
                    QApplication.processEvents()

        except Exception as e:
            QMessageBox.critical(self, "Błąd pobierania",
                                 f"Nie udało się pobrać aktualizacji:\n{e}")
            return
        finally:
            progress.close()

        # Uruchom instalator (Inno Setup) i zamknij aplikację
        subprocess.Popen([tmp_path, "/VERYSILENT", "/CLOSEAPPLICATIONS", "/RESTARTAPPLICATIONS"])
        QApplication.quit()

    # ---------------------------------------------------------------- helpers – General

    def _on_duty_toggled(self, checked: bool):
        self._db.set_setting("show_duty_section", "1" if checked else "0")
        self._db.set_setting("dyzur_highlight_enabled", "1" if checked else "0")
        self._cb_dyzur_color.setChecked(checked)

    def _on_theme_toggled(self, checked: bool):
        self._db.set_setting("theme_mode", "light" if checked else "dark")

    # ---------------------------------------------------------------- helpers – Columns

    def _move_col_up(self):
        row = self._col_list.currentRow()
        if row > 0:
            item = self._col_list.takeItem(row)
            self._col_list.insertItem(row - 1, item)
            self._col_list.setCurrentRow(row - 1)

    def _move_col_down(self):
        row = self._col_list.currentRow()
        if row < self._col_list.count() - 1:
            item = self._col_list.takeItem(row)
            self._col_list.insertItem(row + 1, item)
            self._col_list.setCurrentRow(row + 1)

    def _set_all_checked(self, state: bool):
        for i in range(self._col_list.count()):
            self._col_list.item(i).setCheckState(
                Qt.Checked if state else Qt.Unchecked
            )

    def _apply_columns(self):
        visible, order = set(), []
        for i in range(self._col_list.count()):
            item = self._col_list.item(i)
            attr = item.data(Qt.UserRole)
            order.append(attr)
            if item.checkState() == Qt.Checked:
                visible.add(attr)
        visible.add("id")
        self._visible_set = visible
        self._column_order = ["id"] + order
        self.columns_changed.emit(visible, ["id"] + order)

    def get_column_result(self) -> tuple[set, list]:
        """Zwraca aktualny stan kolumn (po ewentualnym Zastosuj)."""
        return self._visible_set, self._column_order
