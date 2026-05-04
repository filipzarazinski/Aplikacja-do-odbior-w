"""
ui/dict_tab.py
Generyczny widget zakładki słownika: tabela + CRUD + import z Excela.
"""
import io
import logging
import shutil
import tempfile
import os
from typing import Callable, Optional

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QDialog, QFormLayout,
    QLineEdit, QDialogButtonBox, QFileDialog,
    QLabel, QMessageBox,
)
from PySide6.QtCore import Qt, Signal

logger = logging.getLogger(__name__)


class _RowDialog(QDialog):
    """Prosty dialog do dodawania/edytowania wiersza słownika."""

    def __init__(self, labels: list[str], values: list[str] = None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edycja rekordu")
        self.setModal(True)
        self.setMinimumWidth(340)

        layout = QVBoxLayout(self)
        form = QFormLayout()
        form.setSpacing(8)
        self._edits: list[QLineEdit] = []

        for i, label in enumerate(labels):
            edit = QLineEdit()
            edit.setText(values[i] if values and i < len(values) else "")
            form.addRow(label + ":", edit)
            self._edits.append(edit)

        layout.addLayout(form)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def get_values(self) -> list[str]:
        return [e.text().strip() for e in self._edits]


class DictTab(QWidget):
    """
    Generyczny widget do zarządzania słownikiem.

    Parameters
    ----------
    headers : list[str]
        Nagłówki kolumn widocznych w tabeli.
    loader : () -> list[tuple]
        Pobiera rekordy; każdy tuple: (id, val1, val2, ...).
    saver : (values: list[str]) -> None
        Zapisuje nowy lub edytowany rekord; values = [val1, val2, ...] (bez id).
    deleter : (id: int) -> bool
        Usuwa rekord o podanym id.
    excel_sheet : str
        Nazwa arkusza do importu.
    excel_parser : (worksheet) -> list[list[str]]
        Parsuje arkusz i zwraca listę wierszy (bez id).
    add_labels : list[str]
        Etykiety pól w dialogu Dodaj/Edytuj. Domyślnie = headers.
    commit_fn : callable
        Funkcja do zatwierdzania zmian w DB (np. db.commit()).
    """

    data_changed = Signal()

    def __init__(
        self,
        headers: list[str],
        loader: Callable,
        saver: Callable,
        deleter: Callable,
        excel_sheet: str,
        excel_parser: Callable,
        add_labels: list[str] = None,
        commit_fn: Callable = None,
        lazy: bool = False,
        parent=None,
    ):
        super().__init__(parent)
        self._headers = headers
        self._loader = loader
        self._saver = saver
        self._deleter = deleter
        self._excel_sheet = excel_sheet
        self._excel_parser = excel_parser
        self._add_labels = add_labels or headers
        self._commit = commit_fn or (lambda: None)
        self._lazy = lazy
        self._loaded = False
        self._row_ids: list[int] = []
        self._all_records: list = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(4)

        # Search bar
        search_row = QHBoxLayout()
        search_row.setSpacing(6)
        self._search = QLineEdit()
        self._search.setPlaceholderText("🔍  Szukaj…  (kolumna:tekst)")
        self._search.setFixedHeight(24)
        self._search.setClearButtonEnabled(True)
        self._search.setToolTip(
            "Szukaj we wszystkich kolumnach: wpisz tekst\n"
            "Szukaj w konkretnej kolumnie: nazwa_kolumny:tekst\n"
            "Przykład:  ccid:8948  lub  sim:501"
        )
        self._search.textChanged.connect(self._apply_filter)
        search_row.addWidget(self._search, 1)
        self._info_lbl = QLabel()
        self._info_lbl.setStyleSheet("color: #64748b; font-size: 8pt;")
        search_row.addWidget(self._info_lbl)
        layout.addLayout(search_row)

        # Table
        self._table = QTableWidget(self)
        self._table.setColumnCount(len(headers))
        self._table.setHorizontalHeaderLabels(headers)
        self._table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self._table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self._table.setAlternatingRowColors(True)
        self._table.verticalHeader().setVisible(False)
        self._table.verticalHeader().setDefaultSectionSize(20)
        hdr = self._table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.Stretch)
        hdr.setDefaultSectionSize(80)
        self._table.doubleClicked.connect(self._on_edit)
        layout.addWidget(self._table, 1)

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.setSpacing(4)

        btn_add   = QPushButton("＋ Dodaj")
        btn_edit  = QPushButton("✎ Edytuj")
        btn_del   = QPushButton("✕ Usuń")
        btn_clear = QPushButton("⊘ Usuń wszystkie")
        btn_imp   = QPushButton("📥 Importuj z Excel")

        btn_imp.setToolTip(f"Importuj z arkusza: {excel_sheet}")
        btn_imp.setObjectName("btn_primary")


        btn_add.clicked.connect(self._on_add)
        btn_edit.clicked.connect(self._on_edit)
        btn_del.clicked.connect(self._on_delete)
        btn_clear.clicked.connect(self._on_clear_all)
        btn_imp.clicked.connect(self._on_import)

        btn_row.addWidget(btn_add)
        btn_row.addWidget(btn_edit)
        btn_row.addWidget(btn_del)
        btn_row.addWidget(btn_clear)
        btn_row.addStretch()
        btn_row.addWidget(btn_imp)
        layout.addLayout(btn_row)

        if not lazy:
            self.refresh()
        else:
            self._info_lbl.setText("Nie załadowano")

    # ---------------------------------------------------------------- Public

    def ensure_loaded(self):
        """Ładuje dane jeśli jeszcze nie były ładowane (lazy load)."""
        if not self._loaded:
            self.refresh()

    def refresh(self):
        self._loaded = True
        self._all_records = self._loader()
        self._row_ids = []
        n = len(self._all_records)

        self._table.setSortingEnabled(False)
        self._table.setUpdatesEnabled(False)
        self._table.clearContents()
        self._table.setRowCount(n)

        for i, rec in enumerate(self._all_records):
            self._row_ids.append(rec[0])
            vals = [str(v) if v is not None else "" for v in rec[1:]]
            for col, val in enumerate(vals[:len(self._headers)]):
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self._table.setItem(i, col, item)

        self._table.setUpdatesEnabled(True)
        self._apply_filter()

    def _apply_filter(self):
        raw = self._search.text().strip()
        visible = 0
        total = self._table.rowCount()

        # Parsowanie "kolumna:tekst" — dopasowanie nazwy kolumny (częściowe, bez wielkości liter)
        col_idx: int | None = None
        query = raw.lower()
        if ":" in raw:
            col_part, _, val_part = raw.partition(":")
            col_key = col_part.strip().lower()
            for i, hdr in enumerate(self._headers):
                if col_key in hdr.lower():
                    col_idx = i
                    query = val_part.strip().lower()
                    break

        for row in range(total):
            if not query:
                match = True
            elif col_idx is not None:
                cell = self._table.item(row, col_idx)
                match = query in (cell.text().lower() if cell else "")
            else:
                match = any(
                    query in (self._table.item(row, c).text().lower() if self._table.item(row, c) else "")
                    for c in range(self._table.columnCount())
                )
            self._table.setRowHidden(row, not match)
            if match:
                visible += 1

        if query:
            col_label = f" [{self._headers[col_idx]}]" if col_idx is not None else ""
            self._info_lbl.setText(f"Wyniki{col_label}: {visible} / {total}")
        else:
            self._info_lbl.setText(f"Rekordów: {total}")

    # ---------------------------------------------------------------- Slots

    def _on_add(self):
        dlg = _RowDialog(self._add_labels, parent=self)
        if dlg.exec():
            vals = dlg.get_values()
            if any(vals):
                self._saver(vals)
                self._commit()
                self.refresh()
                self.data_changed.emit()

    def _selected_rows(self) -> list[int]:
        """Zwraca posortowaną listę indeksów zaznaczonych wierszy (bez duplikatów)."""
        return sorted({idx.row() for idx in self._table.selectedIndexes()})

    def _on_edit(self):
        rows = self._selected_rows()
        if len(rows) != 1:
            return
        row = rows[0]
        current = [
            self._table.item(row, c).text()
            for c in range(self._table.columnCount())
        ]
        dlg = _RowDialog(self._add_labels, current, parent=self)
        if dlg.exec():
            vals = dlg.get_values()
            if any(vals):
                self._deleter(self._row_ids[row])
                self._saver(vals)
                self._commit()
                self.refresh()
                self.data_changed.emit()

    def _on_delete(self):
        rows = self._selected_rows()
        if not rows:
            return
        if len(rows) == 1:
            name = self._table.item(rows[0], 0).text()
            msg = f"Usunąć '{name}'?"
        else:
            msg = f"Usunąć zaznaczone {len(rows)} rekordy?"
        if QMessageBox.question(
            self, "Usuń rekord", msg,
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        ) == QMessageBox.Yes:
            for row in rows:
                self._deleter(self._row_ids[row])
            self._commit()
            self.refresh()
            self.data_changed.emit()

    def _on_clear_all(self):
        if not self._row_ids:
            return
        count = len(self._row_ids)
        if QMessageBox.question(
            self, "Usuń wszystkie rekordy",
            f"Na pewno usunąć wszystkie {count} rekord(ów)?\nTej operacji nie można cofnąć.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        ) == QMessageBox.Yes:
            for rid in list(self._row_ids):
                self._deleter(rid)
            self._commit()
            self.refresh()
            self.data_changed.emit()

    def _on_import(self):
        path, _ = QFileDialog.getOpenFileName(
            self, f"Wybierz plik Excel (arkusz: {self._excel_sheet})",
            "", "Pliki Excel (*.xlsx *.xlsm *.xls)"
        )
        if not path:
            return
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "Brak biblioteki",
                                "Uruchom: pip install openpyxl")
            return
        try:
            tmp = tempfile.mktemp(suffix=os.path.splitext(path)[1])
            shutil.copy2(path, tmp)
            wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
            if self._excel_sheet not in wb.sheetnames:
                QMessageBox.warning(
                    self, "Brak arkusza",
                    f"Nie znaleziono arkusza '{self._excel_sheet}'.\n"
                    f"Dostepne: {', '.join(wb.sheetnames)}"
                )
                wb.close()
                os.remove(tmp)
                return
            ws = wb[self._excel_sheet]
            rows = self._excel_parser(ws)
            wb.close()
            os.remove(tmp)
        except Exception as e:
            QMessageBox.critical(self, "Błąd odczytu", str(e))
            return

        count = 0
        for vals in rows:
            if any(v for v in vals if v):
                self._saver(vals)
                count += 1
        self._commit()
        self.refresh()
        self.data_changed.emit()
        QMessageBox.information(self, "Import zakonczony",
                                f"Zaimportowano {count} rekordow z arkusza '{self._excel_sheet}'.")
