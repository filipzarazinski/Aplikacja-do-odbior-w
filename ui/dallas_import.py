"""
ui/dallas_import.py
--------------------
Parsowanie załączników z listą pojazdów firmy (xlsx/xls) i eksportów urządzeń
z serwera (csv), oraz dopasowywanie ich po ID rejestratora / Imei.

Czysta logika, bez zależności od Qt - łatwa do przetestowania niezależnie od UI.
"""
import re
from typing import Optional


def _norm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


_SIM_DIGITS_RE = re.compile(r"\d+")


def _normalize_sim(raw: str) -> str:
    """Sprowadza numer SIM do formatu '+48XXXXXXXXX', niezależnie od tego, jak jest
    zapisany w załączniku (bez '+', bez '48', ze spacjami/myślnikami między cyframi).
    Jeśli formatu nie da się jednoznacznie rozpoznać, zwraca oryginał bez zmian."""
    s = (raw or "").strip()
    if not s:
        return ""
    digits = "".join(_SIM_DIGITS_RE.findall(s))
    if len(digits) == 11 and digits.startswith("48"):
        return "+" + digits
    if len(digits) == 9:
        return "+48" + digits
    return s


_TOKEN_RE = re.compile(r"[^a-ząćęłńóśźż0-9]+")


def _has_token(text: str, token: str) -> bool:
    """Sprawdza czy `token` występuje jako samodzielne słowo w nagłówku, nie jako
    fragment innego słowa (np. 'id' nie ma trafiać w 'widok')."""
    return token in _TOKEN_RE.split(text)


# Rozpoznawanie kolumn po nagłówku - lista wariantów nazw używanych w różnych
# załącznikach (różne firmy/flotę nazywają te same kolumny inaczej). Dopisywanie
# nowego wariantu tutaj wystarcza, żeby nowy format pliku zaczął się rozpoznawać
# automatycznie - nie trzeba niczego zmieniać w resztcie kodu.
_PLATE_HINTS = ("rejestracyjny", "rejestracja", "tablica rej")
# UWAGA: samo "rejestrator" jako podstring jest zbyt szerokie - w załącznikach PGE/Tauron
# jest kolumna "Rodzaj (marka, model) zamontowanego w pojeździe rejestratora" (model
# urządzenia, nie ID!), która występuje PRZED właściwą kolumną "ID/Id rejestratora" i była
# przez to błędnie łapana jako ID. "id" jako samodzielny token (patrz _has_token) już
# poprawnie rozróżnia te dwie kolumny, więc "rejestrator" nie jest tu potrzebne.
_DEVICE_ID_HINTS = ("imei", "numer seryjny", "nr seryjny", "s/n", "kod pojazdu")
_SIM_HINTS = ("sim",)


def _find_header_row(rows: list[list]) -> tuple[int, dict]:
    """Szuka wśród pierwszych ~25 wierszy tego, który wygląda jak nagłówek tabeli
    z kolumnami: nr rejestracyjny, ID rejestratora, nr karty SIM. Nie zakłada
    konkretnego wiersza/kolumny/nazwy nagłówka - różni się to między załącznikami
    różnych firm, więc rozpoznawanie opiera się na liście wariantów nazw kolumn
    (patrz _PLATE_HINTS / _DEVICE_ID_HINTS / _SIM_HINTS).
    """
    best_idx, best_map, best_score = -1, {}, 0
    for i, row in enumerate(rows[:25]):
        col_map = {}
        for c, cell in enumerate(row):
            text = _norm(cell).lower()
            if not text:
                continue
            if "device_id" not in col_map and (
                any(h in text for h in _DEVICE_ID_HINTS) or _has_token(text, "id")
            ):
                col_map["device_id"] = c
            if "plate" not in col_map and any(h in text for h in _PLATE_HINTS):
                col_map["plate"] = c
            if "sim" not in col_map and any(h in text for h in _SIM_HINTS):
                col_map["sim"] = c
        if len(col_map) > best_score:
            best_score, best_idx, best_map = len(col_map), i, col_map
    if best_score < 2:
        raise ValueError(
            "Nie rozpoznano formatu pliku - nie znaleziono kolumn 'Nr rejestracyjny', "
            "'ID rejestratora' i 'Nr karty SIM' (ani ich rozpoznawanych wariantów nazw)."
        )
    return best_idx, best_map


def _is_column_index_junk_row(row: list) -> bool:
    """Wykrywa 'śmieciowy' wiersz z numeracją kolumn (1, 2, 3...) występujący
    czasem tuż pod nagłówkiem (zaobserwowane w załączniku Tauron)."""
    non_empty = [(c, v) for c, v in enumerate(row) if _norm(v)]
    if len(non_empty) < 3:
        return False
    for c, v in non_empty:
        try:
            if float(v) != float(c + 1):
                return False
        except (TypeError, ValueError):
            return False
    return True


def parse_attachment(path: str) -> list[dict]:
    """Parsuje załącznik xlsx/xls z listą pojazdów.
    Zwraca listę dict: {'plate', 'device_id', 'sim'}."""
    rows_by_sheet: list[list[list]] = []

    if path.lower().endswith(".xls"):
        import xlrd
        wb = xlrd.open_workbook(path, ignore_workbook_corruption=True)
        for sheet in wb.sheets():
            rows_by_sheet.append([
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ])
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for name in wb.sheetnames:
            rows_by_sheet.append([list(row) for row in wb[name].iter_rows(values_only=True)])
        wb.close()

    result: list[dict] = []
    found_any_sheet = False

    for rows in rows_by_sheet:
        if not rows:
            continue
        try:
            header_idx, col_map = _find_header_row(rows)
        except ValueError:
            continue
        found_any_sheet = True

        def get(row: list, key: str) -> str:
            idx = col_map.get(key)
            return _norm(row[idx]) if idx is not None and idx < len(row) else ""

        for row in rows[header_idx + 1:]:
            if _is_column_index_junk_row(row):
                continue
            plate = get(row, "plate")
            device_id = get(row, "device_id")
            # Surowa wartość z załącznika - specjalnie NIE normalizowana tutaj, żeby błędny
            # format w tabeli był widoczny jako sygnał "trzeba to poprawić w systemie".
            # Normalizacja (patrz _normalize_sim) stosowana jest tylko przy eksporcie do CSV.
            sim = get(row, "sim")
            if not plate and not device_id:
                continue
            result.append({"plate": plate, "device_id": device_id, "sim": sim})

    if not found_any_sheet:
        raise ValueError(
            "Nie rozpoznano formatu pliku - nie znaleziono kolumn 'Nr rejestracyjny', "
            "'ID rejestratora' i 'Nr karty SIM' w żadnym arkuszu."
        )
    return result


# Nazwy kolumn w eksporcie CSV z serwera (identyczne dla różnych firm - różni się
# tylko kolejność/liczba pozostałych kolumn, więc szukamy po nazwie, nie po indeksie).
_SERVER_COLUMNS = {
    "firmware": "Firmware urządzenia",
    "ccrc": "Crcc",
    "firmware_status": "Firmware status",
    "last_gps": "Ostatnia data GPS",
    "dallas": "DALLAS",
    "dallas_error_date": "Data zarejestrowania błędu wysyłania listy dallas",
}
_SERVER_ID_COLUMN = "Imei"
_SERVER_FLEET_COLUMN = "Flota"


def _normalize_fleet_label(raw: str) -> str:
    low = raw.lower()
    if "pge" in low:
        return "PGE"
    if "tauron" in low:
        return "Tauron"
    return raw.strip()


def parse_server_export(path: str) -> tuple[dict[str, dict], str]:
    """Parsuje eksport CSV urządzeń z serwera.
    Zwraca (device_id -> dane, wykryta_flota) - flota to najczęstsza wartość
    kolumny 'Flota' w pliku (np. "PGE"/"Tauron"), albo "" jeśli nie rozpoznano."""
    import csv

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        first_line = f.readline()
        f.seek(0)
        delimiter = ";" if first_line.count(";") >= first_line.count(",") else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        if not reader.fieldnames or _SERVER_ID_COLUMN not in reader.fieldnames:
            raise ValueError(
                f"Nie znaleziono kolumny '{_SERVER_ID_COLUMN}' (ID urządzenia) "
                "w pliku eksportu z serwera."
            )
        result: dict[str, dict] = {}
        fleet_counts: dict[str, int] = {}
        for row in reader:
            device_id = _norm(row.get(_SERVER_ID_COLUMN))
            if not device_id:
                continue
            result[device_id] = {key: _norm(row.get(col)) for key, col in _SERVER_COLUMNS.items()}
            fleet_val = _norm(row.get(_SERVER_FLEET_COLUMN))
            if fleet_val:
                fleet_counts[fleet_val] = fleet_counts.get(fleet_val, 0) + 1

    detected_fleet = ""
    if fleet_counts:
        most_common = max(fleet_counts, key=fleet_counts.get)
        detected_fleet = _normalize_fleet_label(most_common)

    return result, detected_fleet


def build_matched_rows(attachment_rows: list[dict], server_map: dict[str, dict]) -> list[dict]:
    """Łączy wiersze załącznika z danymi serwera po device_id (dokładny match stringowy)."""
    rows = []
    for a in attachment_rows:
        srv: Optional[dict] = server_map.get(a["device_id"]) if a["device_id"] else None
        rows.append({
            "plate": a["plate"],
            "device_id": a["device_id"],
            "sim": a["sim"],
            "found": srv is not None,
            "firmware": srv.get("firmware", "") if srv else "",
            "ccrc": srv.get("ccrc", "") if srv else "",
            "firmware_status": srv.get("firmware_status", "") if srv else "",
            "last_gps": srv.get("last_gps", "") if srv else "",
            "dallas": srv.get("dallas", "") if srv else "",
            "dallas_error_date": srv.get("dallas_error_date", "") if srv else "",
        })
    return rows
